#!/usr/bin/env python3

from __future__ import annotations
from error_handler import handle_exception

from collections import Counter
import argparse
import csv
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
Path("output").mkdir(exist_ok=True)
from typing import Iterable, Iterator, Optional, Set, Tuple
from console import *


def localname(tag: str) -> str:
    return tag.split("}", 1)[1] if "}" in tag else tag


def find_first_desc_text(elem: ET.Element, wanted_names: Iterable[str]) -> Optional[str]:
    wanted = set(wanted_names)
    for d in elem.iter():
        if localname(d.tag) in wanted and d.text and d.text.strip():
            return d.text.strip()
    return None


def find_first_sourcelocation(elem: ET.Element) -> Tuple[Optional[str], Optional[str]]:
    path_keys = ("path", "filepath", "file", "sourcePath")
    line_keys = ("line", "lineStart", "startLine", "lineNumber")

    for d in elem.iter():
        if localname(d.tag) in ("SourceLocation", "Location", "FileLocation"):
            attrs = d.attrib or {}

            path = next(
                (
                    attrs[k].strip()
                    for k in path_keys
                    if k in attrs and attrs[k].strip()
                ),
                None,
            )

            line = next(
                (
                    attrs[k].strip()
                    for k in line_keys
                    if k in attrs and attrs[k].strip()
                ),
                None,
            )

            if path or line:
                return path, line

    return None, None


def get_instance_id_from_vuln(vuln_elem: ET.Element) -> Optional[str]:
    for d in vuln_elem.iter():
        if localname(d.tag) == "InstanceID" and d.text and d.text.strip():
            return d.text.strip()
    return None


def build_vulnerability_name(vuln_elem: ET.Element) -> str:
    vuln_type = find_first_desc_text(vuln_elem, ("Type",))
    vuln_subtype = find_first_desc_text(vuln_elem, ("Subtype",))

    if vuln_type and vuln_subtype:
        return f"{vuln_type}: {vuln_subtype}"

    if vuln_subtype:
        return vuln_subtype

    if vuln_type:
        return vuln_type

    return (
        vuln_elem.attrib.get("classID")
        or vuln_elem.attrib.get("ClassID")
        or "Unknown"
    )
    
def get_severity_from_vuln(vuln_elem: ET.Element) -> str:
    severity = ""

    for tag in ("Severity", "DefaultSeverity", "Friority"):
        value = find_first_desc_text(vuln_elem, (tag,))
        if value:
            severity = value.strip()
            break

    severity_map = {
        "5": "Critical",
        "5.0": "Critical",
        "4": "High",
        "4.0": "High",
        "3": "Medium",
        "3.0": "Medium",
        "2": "Low",
        "2.0": "Low",
        "1": "Info",
        "1.0": "Info"
    }

    return severity_map.get(severity, severity)


def load_suppressed_instance_ids_from_audit_xml(zf: zipfile.ZipFile) -> Set[str]:
    suppressed: Set[str] = set()

    candidates = [
        n for n in zf.namelist()
        if n.lower().endswith("audit.xml")
    ]

    if not candidates:
        return suppressed

    audit_name = candidates[0]

    with zf.open(audit_name, "r") as f:
        for event, elem in ET.iterparse(f, events=("end",)):
            if localname(elem.tag) == "Issue":
                inst = (
                    elem.attrib.get("instanceId")
                    or elem.attrib.get("instanceID")
                )

                sup = (
                    elem.attrib.get("suppressed")
                    or elem.attrib.get("Suppressed")
                )

                if inst and sup and sup.strip().lower() == "true":
                    suppressed.add(inst.strip())

                elem.clear()

    return suppressed


def iter_unsuppressed_from_fpr(fpr_path: str) -> Iterator[Tuple[str, str, str]]:
    with zipfile.ZipFile(fpr_path, "r") as zf:

        suppressed_ids = load_suppressed_instance_ids_from_audit_xml(zf)

        fvdl_candidates = [
            n for n in zf.namelist()
            if n.lower().endswith("audit.fvdl")
        ]

        if not fvdl_candidates:
            raise FileNotFoundError(
                "No audit.fvdl found inside the .fpr archive."
            )

        fvdl_name = fvdl_candidates[0]

        with zf.open(fvdl_name, "r") as f:

            for event, elem in ET.iterparse(f, events=("end",)):

                if localname(elem.tag) == "Vulnerability":

                    instance_id = get_instance_id_from_vuln(elem)

                    if instance_id and instance_id in suppressed_ids:
                        elem.clear()
                        continue

                    vuln_name = build_vulnerability_name(elem)

                    path, line = find_first_sourcelocation(elem)

                    if path and line:
                        file_line = f"{path}, line {line}"
                    elif path:
                        file_line = f"{path}, line "
                    elif line:
                        file_line = f", line {line}"
                    else:
                        file_line = ""

                    severity = get_severity_from_vuln(elem)
                    yield vuln_name, file_line, severity

                    elem.clear()


def default_output_path(fpr_path: str) -> str:
    p = Path(fpr_path)
    return str(p.with_suffix(".csv"))
    
def iter_all_findings(fpr_files):
    for fpr in fpr_files:
        for vuln_name, file_line, severity in iter_unsuppressed_from_fpr(fpr):
            yield vuln_name, file_line, severity

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Extract unsuppressed Fortify findings from .fpr"
    )

    ap.add_argument(
    "fprs",
    nargs="+",
    help="One or more FPR files"
    )

    ap.add_argument(
        "-o",
        "--out",
        default=None,
        help="Output Excel path"
    )

    args = ap.parse_args()

    from collections import defaultdict

    out_path = args.out if args.out else "Fortify_Findings.xlsx"

    try:
        findings = list(iter_all_findings(args.fprs))

        grouped = defaultdict(list)

        for vuln_name, endpoint, severity in findings:
            grouped[vuln_name].append((endpoint, severity))

        wb = Workbook()
        ws = wb.active
        ws.title = "Fortify Findings"

        headers = [
            "S. No.",
            "Vulnerability Name",
            "Occurrences",
            "Endpoints",
            "Severity"
        ]

        header_font = Font(
            name="Calibri",
            size=12,
            bold=True,
            color="FFFFFF"
        )

        data_font = Font(
            name="Calibri",
            size=12
        )

        header_fill = PatternFill(
            fill_type="solid",
            start_color="4472C4",
            end_color="4472C4"
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        serial_no = 1
        excel_row = 2

        SEVERITY_ORDER = {
            "Critical": 1,
            "High": 2,
            "Medium": 3,
            "Low": 4,
            "Info": 5
        }

        sorted_vulns = sorted(
            grouped.items(),
            key=lambda item: (
                SEVERITY_ORDER.get(
                    item[1][0][1],
                    999
                ),
                item[0]
            )
        )

        for vuln_name, entries in sorted_vulns:

            endpoints = [x[0] for x in entries]
            severity = entries[0][1]

            ws.cell(excel_row, 1, serial_no)
            ws.cell(excel_row, 2, vuln_name)
            ws.cell(excel_row, 3, len(endpoints))
            ws.cell(excel_row, 4, endpoints[0])
            ws.cell(excel_row, 5, severity)

            excel_row += 1

            for endpoint in endpoints[1:]:
                ws.cell(excel_row, 4, endpoint)
                ws.cell(excel_row, 5, severity)
                excel_row += 1

            serial_no += 1

        for row in ws.iter_rows():
            for cell in row:

                if cell.row > 1:
                    cell.font = data_font

                cell.border = thin_border

                if cell.column in (1, 3, 5):
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )
                else:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center"
                    )

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 100
        ws.column_dimensions["E"].width = 15

        ws.freeze_panes = "A2"

        xlsx_path = out_path
        if not xlsx_path.lower().endswith(".xlsx"):
            xlsx_path += ".xlsx"


        critical_fill = PatternFill("solid", fgColor="C00000")
        high_fill = PatternFill("solid", fgColor="FF0000")
        medium_fill = PatternFill("solid", fgColor="FFC000")
        low_fill = PatternFill("solid", fgColor="00B0F0")
        info_fill = PatternFill("solid", fgColor="92D050")

        severity_font = Font(
            name="Calibri",
            size=12,
            bold=True,
            color="FFFFFF"
        )

        for row in range(2, ws.max_row + 1):

            cell = ws.cell(row=row, column=5)

            if not cell.value:
                continue

            severity = str(cell.value).strip().lower()

            if severity == "critical":
                cell.fill = critical_fill
                cell.font = severity_font

            elif severity == "high":
                cell.fill = high_fill
                cell.font = severity_font

            elif severity == "medium":
                cell.fill = medium_fill
                cell.font = severity_font

            elif severity == "low":
                cell.fill = low_fill
                cell.font = severity_font
            elif severity == "info":
                cell.fill = info_fill
                cell.font = severity_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
        wb.save("output\\Fortify_Findings_Parser.xlsx")

        total_findings = sum(len(v) for v in grouped.values())

        success(
            f"Wrote {total_findings} findings from {len(args.fprs)} FPR file(s) to Fortify_Findings_Parser.xlsx"
        )

        return 0

    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2


def run():
    main()
    
    

if __name__ == "__main__":

    try:

        main()

    except Exception as e:

        handle_exception(e)