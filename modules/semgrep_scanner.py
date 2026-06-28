#!/usr/bin/env python3

import json
import subprocess
from collections import defaultdict
import sys
from pathlib import Path
from console import *

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


RULE_FOLDER = (
    Path(__file__).resolve().parent.parent
    / "rules"
    / "semgrep_rules"
)
Path("output").mkdir(exist_ok=True)

JSON_OUTPUT = "output\\Semgrep_Integration.json"
EXCEL_OUTPUT = "output\\Semgrep_Integration_Findings.xlsx"


def run_semgrep(source_folder):

    cmd = [
        "semgrep",
        "--config",
        str(RULE_FOLDER),
        source_folder,
        "--json",
        "-o",
        JSON_OUTPUT,
        "--max-target-bytes",
        "0",
        "--timeout",
        "120",
        "--timeout-threshold",
        "20",
        "--jobs",
        "4",
        "--no-git-ignore"
    ]

    print("Running Semgrep...")

    subprocess.run(cmd, check=True)


def load_findings():

    with open(JSON_OUTPUT, encoding="utf-8") as f:
        data = json.load(f)

    findings = defaultdict(list)

    severity_map = {
        "ERROR": "High",
        "WARNING": "Medium",
        "INFO": "Low"
    }

    for result in data.get("results", []):

        vuln_name = result["extra"]["message"]

        path = result["path"]

        line = result["start"]["line"]

        severity = severity_map.get(
            result["extra"]["severity"],
            "Medium"
        )

        findings[vuln_name].append(
            (
                path,
                line,
                severity
            )
        )

    return findings


def create_excel(findings):

    wb = Workbook()
    ws = wb.active

    ws.title = "Semgrep Findings"

    headers = [
        "S. No.",
        "Vulnerability Name",
        "Occurrences",
        "Endpoints",
        "Severity"
    ]

    header_font = Font(
        name="Calibri",
        size=12,
        bold=True,
        color="FFFFFF"
    )

    data_font = Font(
        name="Calibri",
        size=12
    )

    header_fill = PatternFill(
        fill_type="solid",
        start_color="4472C4",
        end_color="4472C4"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(1, col)

        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    row = 2
    serial = 1

    for vuln_name in sorted(findings.keys()):

        entries = findings[vuln_name]

        first = True

        for path, line_no, severity in entries:

            endpoint = f"{path}, line {line_no}"

            if first:

                ws.cell(row, 1, serial)
                ws.cell(row, 2, vuln_name)
                ws.cell(row, 3, len(entries))
                ws.cell(row, 4, endpoint)
                ws.cell(row, 5, severity)

                serial += 1
                first = False

            else:

                ws.cell(row, 4, endpoint)
                ws.cell(row, 5, severity)

            row += 1

    severity_colors = {
        "Critical": "C00000",
        "High": "FF0000",
        "Medium": "FFC000",
        "Low": "00B0F0",
        "Info": "92D050"
    }

    for r in ws.iter_rows():

        for cell in r:

            if cell.row > 1:
                cell.font = data_font

            cell.border = thin_border

            if cell.column in (1, 3, 5):

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            else:

                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center"
                )

    for r in range(2, ws.max_row + 1):

        cell = ws.cell(r, 5)

        if not cell.value:
            continue

        sev = str(cell.value)

        if sev in severity_colors:

            cell.fill = PatternFill(
                fill_type="solid",
                start_color=severity_colors[sev],
                end_color=severity_colors[sev]
            )

            cell.font = Font(
                name="Calibri",
                size=12,
                bold=True,
                color="FFFFFF"
            )

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 120
    ws.column_dimensions["E"].width = 15

    ws.freeze_panes = "A2"

    wb.save(EXCEL_OUTPUT)

    success(f"{EXCEL_OUTPUT} generated successfully")


def main():

    if len(sys.argv) != 2:
        print("Usage: python semgrep_scanner.py <source_folder>")
        sys.exit(1)

    source_folder = sys.argv[1]

    run_semgrep(source_folder)

    findings = load_findings()

    create_excel(findings)


def run():
    main()
    
from error_handler import handle_exception

if __name__ == "__main__":

    try:

        main()

    except Exception as e:

        handle_exception(e)