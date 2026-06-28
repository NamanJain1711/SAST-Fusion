#!/usr/bin/env python3

import json
import re
import sys
from pathlib import Path
from console import success, output
from console import *

Path("output").mkdir(exist_ok=True)
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


IGNORE_DIRS = {
    "node_modules",
    ".git",
    ".svn",
    ".hg",
    ".idea",
    ".vscode",
    "__pycache__"
}

IGNORE_EXTENSIONS = {
    ".exe",
    ".dll",
    ".so",
    ".jar",
    ".war",
    ".ear",
    ".zip",
    ".rar",
    ".7z",
    ".tar",
    ".gz",
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".bmp",
    ".ico",
    ".pdf",
    ".mp3",
    ".mp4",
    ".avi",
    ".mov",
    ".woff",
    ".woff2",
    ".ttf",
    ".eot",
    ".class",
    ".pyc",
    ".pdb"
}


def load_rules():

    rules_file = (
        Path(__file__).resolve().parent.parent
        / "rules"
        / "rules_v1.json"
    )

    with open(
        rules_file,
        "r",
        encoding="utf-8"
    ) as f:

        return json.load(f)


def should_scan(path):

    for part in path.parts:
        if part.lower() in IGNORE_DIRS:
            return False

    if path.suffix.lower() in IGNORE_EXTENSIONS:
        return False

    return True


def scan_source(root_folder, rules):

    findings = defaultdict(list)

    for file_path in Path(root_folder).rglob("*"):

        if not file_path.is_file():
            continue

        if not should_scan(file_path):
            continue

        try:
            with open(
                file_path,
                "r",
                encoding="utf-8",
                errors="ignore"
            ) as f:

                lines = f.readlines()

        except Exception:
            continue

        for line_no, line in enumerate(lines, start=1):

            for vuln_name, rule in rules.items():

                if re.search(
                    rule["regex"],
                    line,
                    re.IGNORECASE
                ):

                    findings[vuln_name].append(
                        (
                            str(file_path),
                            line_no,
                            rule["severity"]
                        )
                    )

    return findings


def create_excel(findings):

    wb = Workbook()
    ws = wb.active
    ws.title = "Manual SAST Findings"

    headers = [
        "S. No.",
        "Vulnerability Name",
        "Occurrences",
        "Endpoints",
        "Severity"
    ]

    header_font = Font(
        name="Calibri",
        size=12,
        bold=True,
        color="FFFFFF"
    )

    data_font = Font(
        name="Calibri",
        size=12
    )

    header_fill = PatternFill(
        fill_type="solid",
        start_color="4472C4",
        end_color="4472C4"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(1, col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    row = 2
    serial = 1

    for vuln_name in sorted(findings.keys()):

        entries = findings[vuln_name]

        first = True

        for file_path, line_no, severity in entries:

            endpoint = f"{file_path}, line {line_no}"

            if first:

                ws.cell(row, 1, serial)
                ws.cell(row, 2, vuln_name)
                ws.cell(row, 3, len(entries))
                ws.cell(row, 4, endpoint)
                ws.cell(row, 5, severity)

                first = False
                serial += 1

            else:

                ws.cell(row, 4, endpoint)
                ws.cell(row, 5, severity)

            row += 1

    severity_colors = {
        "Critical": "C00000",
        "High": "FF0000",
        "Medium": "FFC000",
        "Low": "00B0F0",
        "Info": "92D050"
    }

    for r in ws.iter_rows():

        for cell in r:

            if cell.row > 1:
                cell.font = data_font

            cell.border = thin_border

            if cell.column in (1, 3, 5):
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
            else:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center"
                )

    for r in range(2, ws.max_row + 1):

        cell = ws.cell(r, 5)

        if not cell.value:
            continue

        sev = str(cell.value)

        if sev in severity_colors:

            cell.fill = PatternFill(
                fill_type="solid",
                start_color=severity_colors[sev],
                end_color=severity_colors[sev]
            )

            cell.font = Font(
                name="Calibri",
                size=12,
                bold=True,
                color="FFFFFF"
            )

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 120
    ws.column_dimensions["E"].width = 15

    ws.freeze_panes = "A2"

    wb.save("output\\Custom_Rules_Engine.xlsx")


def main():

    if len(sys.argv) != 2:
        print("Usage: python scanner.py <source_code_folder>")
        sys.exit(1)

    source_folder = sys.argv[1]

    rules = load_rules()

    findings = scan_source(
        source_folder,
        rules
    )

    create_excel(findings)

    success("Custom_Rules_Engine.xlsx generated successfully")

def run():
    main()
    
from error_handler import handle_exception

if __name__ == "__main__":

    try:

        main()

    except Exception as e:

        handle_exception(e)