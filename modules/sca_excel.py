#!/usr/bin/env python3

from collections import Counter

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

SEVERITY_COLORS = {

    "Critical":"C00000",

    "High":"FF0000",

    "Medium":"FFC000",

    "Low":"00B0F0",

    "Unknown":"808080"

}

HEADER_FILL = PatternFill(

    fill_type="solid",

    start_color="4472C4",

    end_color="4472C4"

)

HEADER_FONT = Font(

    name="Calibri",

    size=12,

    bold=True,

    color="FFFFFF"

)

DATA_FONT = Font(

    name="Calibri",

    size=12

)

BORDER = Border(

    left=Side(style="thin"),

    right=Side(style="thin"),

    top=Side(style="thin"),

    bottom=Side(style="thin")

)

def create_sca_excel(findings):

    severity_order = {
            "Critical": 1,
            "High": 2,
            "Medium": 3,
            "Low": 4,
            "Unknown": 5
        }

    findings = sorted(
        findings,
        key=lambda x: (
            severity_order.get(x["severity"], 99),
            x["package"],
            x["cve"]
        )
    )

    wb = Workbook()

    ws = wb.active

    ws.title = "SCA Findings"

    headers = [

    "S.No",

    "Package",

    "Version",

    "CVE",

    "Occurrences",

    "Dependency Files",

    "Severity",

    "Fixed Version",

    "Scanner"

    ]

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(1, col)

        cell.value = header

        cell.font = HEADER_FONT

        cell.fill = HEADER_FILL

        cell.border = BORDER

        if cell.column == 6:

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        elif cell.column in (1,5,7):

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        else:

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )



    serial = 1

    row = 2

    for finding in findings:

        ws.cell(row=row, column=1, value=serial)

        ws.cell(row,2,finding["package"])

        ws.cell(row,3,finding["version"])

        ws.cell(row,4,finding["cve"])

        ws.cell(row,5,finding["occurrences"])

        ws.cell(
            row,
            6,
            "\n".join(
                finding["dependency_files"]
            )
        )

        ws.cell(
            row,
            7,
            finding["severity"]
        )

        ws.cell(
            row,
            8,
            finding["fixed_version"] or "N/A"
        )

        ws.cell(
            row,
            9,
            finding["scanner"] or "Unknown"
        )

        ws.row_dimensions[row].height = 35

        serial += 1

        row += 1

    for r in ws.iter_rows():

        for cell in r:

            cell.border = BORDER

            if cell.row > 1:

                cell.font = DATA_FONT

                if cell.column == 6:

                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="top",
                        wrap_text=True
                    )

                elif cell.column in (1,5,7):

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

                else:

                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center"
                    )

                
    for r in range(2, ws.max_row + 1):

        cell = ws.cell(r, 7)

        sev = str(cell.value)

        if sev in SEVERITY_COLORS:

            cell.fill = PatternFill(
                fill_type="solid",
                start_color=SEVERITY_COLORS[sev],
                end_color=SEVERITY_COLORS[sev]
            )

            cell.font = Font(
                name="Calibri",
                size=12,
                bold=True,
                color="FFFFFF"
            )

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 70
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 15

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = ws.dimensions

    

    summary = wb.create_sheet("Summary")

    severity_count = Counter(
    finding["severity"]
    for finding in findings
)

    summary["A1"] = "Metric"
    summary["B1"] = "Value"

    summary["A2"] = "Total Findings"
    summary["B2"] = len(findings)

    summary["A3"] = "Unique Packages"
    summary["B3"] = len(
        set(
            f["package"]
            for f in findings
        )
    )

    summary["A4"] = "Critical"
    summary["B4"] = severity_count.get(
        "Critical",
        0
    )

    summary["A5"] = "High"
    summary["B5"] = severity_count.get(
        "High",
        0
    )

    summary["A6"] = "Medium"
    summary["B6"] = severity_count.get(
        "Medium",
        0
    )

    summary["A7"] = "Low"
    summary["B7"] = severity_count.get(
        "Low",
        0
    )

    summary.freeze_panes = "A2"

    for cell in summary[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for row in summary.iter_rows():
        for cell in row:
            cell.border = BORDER

            if cell.row > 1:
                cell.font = DATA_FONT

    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 20

    wb.save(
            "output\\SCA_Findings.xlsx"
        )
    