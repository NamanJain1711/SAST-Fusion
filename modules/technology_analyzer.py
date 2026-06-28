#!/usr/bin/env python3

import sys
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

Path("output").mkdir(exist_ok=True)

IGNORE_DIRS = {
    "node_modules",
    ".git",
    ".svn",
    ".hg",
    ".idea",
    ".vscode",
    "__pycache__"
}

IGNORE_EXTENSIONS = {
    ".exe", ".dll", ".so", ".jar", ".war", ".ear",
    ".zip", ".rar", ".7z", ".tar", ".gz",
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".ico",
    ".pdf", ".mp3", ".mp4", ".avi", ".mov",
    ".woff", ".woff2", ".ttf", ".eot",
    ".class", ".pyc", ".pdb"
}

TECH_MAP = {
    ".java": "Java",
    ".js": "JavaScript",
    ".ts": "TypeScript",
    ".jsx": "React",
    ".tsx": "React TypeScript",
    ".cs": "C#",
    ".vb": "VB.NET",
    ".php": "PHP",
    ".py": "Python",
    ".go": "Go",
    ".xml": "XML",
    ".html": "HTML",
    ".css": "CSS",
    ".sql": "SQL",
    ".json": "JSON",
    ".properties": "Properties",
    ".yml": "YAML",
    ".yaml": "YAML"
}


def should_scan(path):

    for part in path.parts:
        if part.lower() in IGNORE_DIRS:
            return False

    if path.suffix.lower() in IGNORE_EXTENSIONS:
        return False

    return True


def analyze_source(root_folder):

    tech_stats = defaultdict(
        lambda: {
            "files": 0,
            "loc": 0
        }
    )

    file_stats = []

    total_files = 0
    total_loc = 0
    code_loc = 0
    comment_loc = 0
    blank_loc = 0

    for file_path in Path(root_folder).rglob("*"):

        if not file_path.is_file():
            continue

        if not should_scan(file_path):
            continue

        ext = file_path.suffix.lower()

        if ext.replace(".", "").isdigit():
            continue

        technology = TECH_MAP.get(
            ext,
            f"Other ({ext})"
        )

        try:

            with open(
                file_path,
                "r",
                encoding="utf-8",
                errors="ignore"
            ) as f:

                lines = f.readlines()

        except Exception:
            continue

        total_files += 1

        file_loc = 0

        for line in lines:

            stripped = line.strip()

            total_loc += 1

            if not stripped:
                blank_loc += 1
                continue

            if (
                stripped.startswith("//")
                or stripped.startswith("#")
                or stripped.startswith("/*")
                or stripped.startswith("*")
            ):
                comment_loc += 1
                continue

            code_loc += 1
            file_loc += 1

        tech_stats[(technology, ext)]["files"] += 1
        tech_stats[(technology, ext)]["loc"] += file_loc

        
        file_stats.append(
            (
                str(file_path),
                technology,
                ext,
                file_loc
            )
        )

    return (
        tech_stats,
        file_stats,
        total_files,
        total_loc,
        code_loc,
        comment_loc,
        blank_loc
    )


def create_excel(
    tech_stats,
    file_stats,
    total_files,
    total_loc,
    code_loc,
    comment_loc,
    blank_loc
):

    wb = Workbook()

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="4472C4"
    )

    border = Border()

    # Sheet 1

    ws1 = wb.active
    ws1.title = "Technology Summary"

    headers = [
        "Technology",
        "Extension",
        "Files",
        "LOC"
    ]

    for col, header in enumerate(headers, start=1):

        cell = ws1.cell(1, col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    row = 2

    for (technology, ext), data in sorted(
        tech_stats.items(),
        key=lambda x: x[1]["loc"],
        reverse=True
    ):

        ws1.cell(row, 1, technology)
        ws1.cell(row, 2, ext)
        ws1.cell(row, 3, data["files"])
        ws1.cell(row, 4, data["loc"])

        row += 1

    # Sheet 2

    ws2 = wb.create_sheet(
        "Application Metrics"
    )

    metrics = [
        ("Total Files", total_files),
        ("Total LOC", total_loc),
        ("Code LOC", code_loc),
        ("Comment LOC", comment_loc),
        ("Blank LOC", blank_loc)
    ]

    for row, metric in enumerate(metrics, start=1):

        ws2.cell(row, 1, metric[0])
        ws2.cell(row, 2, metric[1])

    # Sheet 3

    ws3 = wb.create_sheet(
        "Largest Files"
    )

    headers = [
        "File",
        "Technology",
        "Extension",
        "LOC"
    ]


    for col, header in enumerate(headers, start=1):

        cell = ws3.cell(1, col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    row = 2

    for file_path, technology, ext, loc in sorted(
        file_stats,
        key=lambda x: x[3],
        reverse=True
    )[:5]:

        ws3.cell(row, 1, file_path)
        ws3.cell(row, 2, technology)
        ws3.cell(row, 3, ext)
        ws3.cell(row, 4, loc)

        row += 1

    wb.save(
        "output\\Technology_Discovery.xlsx"
    )

    success(
        "Technology_Discovery.xlsx generated successfully"
    )


def main():

    if len(sys.argv) != 2:
        print(
            "Usage: python technology_analyzer.py <source_folder>"
        )
        sys.exit(1)

    source_folder = sys.argv[1]

    (
        tech_stats,
        file_stats,
        total_files,
        total_loc,
        code_loc,
        comment_loc,
        blank_loc
    ) = analyze_source(source_folder)

    create_excel(
        tech_stats,
        file_stats,
        total_files,
        total_loc,
        code_loc,
        comment_loc,
        blank_loc
    )


def run():
    main()


from error_handler import handle_exception

if __name__ == "__main__":

    try:

        main()

    except Exception as e:

        handle_exception(e)