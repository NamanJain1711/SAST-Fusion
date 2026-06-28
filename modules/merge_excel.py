#!/usr/bin/env python3

import sys
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from console import *


SEVERITY_ORDER = {
    "Critical": 1,
    "High": 2,
    "Medium": 3,
    "Low": 4,
    "Info": 5
}


from pathlib import Path


def load_report(path):

    ext = Path(path).suffix.lower()

    if ext == ".csv":

        df = pd.read_csv(path)

    elif ext in (".xlsx", ".xls"):

        df = pd.read_excel(path)

    else:

        raise ValueError(
            f"Unsupported file format: {ext}"
        )

    df = df.fillna("")

    return df


def merge_findings(df):

    findings = defaultdict(
        lambda: {
            "severity": "",
            "endpoints": set()
        }
    )

    current_vuln = ""
    current_severity = ""

    for _, row in df.iterrows():

        vuln = str(
            row.get(
                "Vulnerability Name",
                ""
            )
        ).strip()

        sev = str(
            row.get(
                "Severity",
                ""
            )
        ).strip()

        endpoint = str(
            row.get(
                "Endpoints",
                ""
            )
        ).strip()

        if vuln:
            current_vuln = vuln

        if sev:
            current_severity = sev

        if not current_vuln:
            continue

        key = (
            current_vuln,
            current_severity
        )

        findings[key]["severity"] = current_severity

        if endpoint:
            findings[key]["endpoints"].add(
                endpoint
            )

    return findings


def create_excel(findings, output_file):

    wb = Workbook()

    ws = wb.active

    ws.title = "Combined Findings"

    headers = [
        "S. No.",
        "Vulnerability Name",
        "Occurrences",
        "Endpoints",
        "Severity"
    ]

    header_font = Font(
        name="Calibri",
        size=12,
        bold=True,
        color="FFFFFF"
    )

    data_font = Font(
        name="Calibri",
        size=12
    )

    header_fill = PatternFill(
        fill_type="solid",
        start_color="4472C4",
        end_color="4472C4"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, start=1):

        cell = ws.cell(1, col)

        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    sorted_findings = sorted(
        findings.items(),
        key=lambda x: (
            SEVERITY_ORDER.get(
                x[1]["severity"],
                999
            ),
            x[0][0]
        )
    )

    row = 2
    serial = 1

    for (
        vuln_name,
        severity
    ), data in sorted_findings:

        endpoints = sorted(
            data["endpoints"]
        )

        occurrences = len(
            endpoints
        )

        first = True

        for endpoint in endpoints:

            if first:

                ws.cell(
                    row,
                    1,
                    serial
                )

                ws.cell(
                    row,
                    2,
                    vuln_name
                )

                ws.cell(
                    row,
                    3,
                    occurrences
                )

                ws.cell(
                    row,
                    4,
                    endpoint
                )

                ws.cell(
                    row,
                    5,
                    severity
                )

                serial += 1

                first = False

            else:

                ws.cell(
                    row,
                    4,
                    endpoint
                )

                ws.cell(
                    row,
                    5,
                    severity
                )

            row += 1

    severity_colors = {
        "Critical": "C00000",
        "High": "FF0000",
        "Medium": "FFC000",
        "Low": "00B0F0",
        "Info": "92D050"
    }

    for r in ws.iter_rows():

        for cell in r:

            if cell.row > 1:
                cell.font = data_font

            cell.border = thin_border

            if cell.column in (
                1,
                3,
                5
            ):

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            else:

                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center"
                )

    for r in range(
        2,
        ws.max_row + 1
    ):

        cell = ws.cell(
            r,
            5
        )

        sev = str(
            cell.value
        )

        if sev in severity_colors:

            cell.fill = PatternFill(
                fill_type="solid",
                start_color=severity_colors[
                    sev
                ],
                end_color=severity_colors[
                    sev
                ]
            )

            cell.font = Font(
                name="Calibri",
                size=12,
                bold=True,
                color="FFFFFF"
            )

    ws.column_dimensions[
        "A"
    ].width = 10

    ws.column_dimensions[
        "B"
    ].width = 50

    ws.column_dimensions[
        "C"
    ].width = 15

    ws.column_dimensions[
        "D"
    ].width = 120

    ws.column_dimensions[
        "E"
    ].width = 15

    ws.freeze_panes = "A2"

    import os

    os.makedirs("output", exist_ok=True)

    if not output_file.lower().endswith(".xlsx"):
        output_file += ".xlsx"

    output_path = os.path.join(
        "output",
        output_file
    )

    wb.save(output_path)

    return output_path


def main():

    if len(sys.argv) != 4:

        print(
            "Usage: python merge_excel_v2.py fortify.xlsx manual.xlsx output.xlsx"
        )

        sys.exit(1)

    df1 = load_report(sys.argv[1])
    df2 = load_report(sys.argv[2])

    merged_df = pd.concat(
        [df1, df2],
        ignore_index=True
    )

    findings = merge_findings(
        merged_df
    )

    output_path = create_excel(
        findings,
        sys.argv[3]
    )

    print("\n")
    print("=" * 50)
    print("Report Consolidator")
    print("=" * 50)
    success("✔ Reports merged successfully.")
    success(f"Output : {output_path}")

def run():
    main()

from error_handler import handle_exception

if __name__ == "__main__":

    try:

        main()

    except Exception as e:

        handle_exception(e)